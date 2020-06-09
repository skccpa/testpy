import openpyxl

from pathlib import Path

#Dev tools path
#Path(r"C:\Users\Kyle\CKC Advisors\CKC Sharept Dev - Documents\Sharepoint Project")
#print("dev tools path: " + devTools_path)

#Lookup from/source WB
source_wb = (r'C:\Users\Kyle\CKC Advisors\CKC Sharept Dev - Documents\Sharepoint Project\Lists_Live\DataMap_Excel.xlsx')
wk = openpyxl.load_workbook(source_wb)
print(wk.sheetnames)

#access sheet
sh_dm = wk['DM_PBI']

#access shell/range of cells
cell = sh_dm['A9']
print(cell.value)

#get row
print(cell.row)
print(cell.column)
print(cell.coordinate)
print(sh_dm.max_row + sh_dm.max_column)

#for loop with individual cells
#for row in range(1, sh_dm.max_row + 1):
    #for column in range(1, sh_dm.max_column + 1):
    #    cell = sh_dm.cell(row, column)
    #    print(cell.value)


#range of cells - all cells in col A
#column = sh_dm["a"]
#print(column)

#range of columns
#cells = sh_dm(["a:c"])

print(sh_dm["a9:c20"])




        



